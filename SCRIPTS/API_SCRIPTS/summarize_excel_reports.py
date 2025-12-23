#!/usr/bin/env python3
"""
Script to summarize pass/fail status from all Excel files in PythonWorkingScripts_Output directory.
The 2nd column of each Excel file contains pass/fail status.
Creates a summary Excel file with failure counts out of total for each Excel file.
"""

import os
import xlrd
import xlsxwriter
import shutil
import tempfile
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. .xlsx files may not be readable.")


def find_all_excel_files(directory):
    """Find all .xls and .xlsx files recursively in the given directory."""
    excel_files = []
    for root, dirs, files in os.walk(directory):
        # Skip hidden directories and lock files
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        for file in files:
            if file.endswith(('.xls', '.xlsx')) and not file.startswith('.~lock') and not file.startswith('Excel_Reports_Summary'):
                full_path = os.path.join(root, file)
                excel_files.append(full_path)
    return excel_files


def parse_datetime_string(dt_str):
    """Parse datetime string from Excel and return datetime object."""
    if not dt_str or dt_str == '':
        return None
    dt_str = str(dt_str).strip()
    # Try common datetime formats
    formats = [
        "%Y-%m-%d-%H-%M-%S",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(dt_str, fmt)
        except:
            continue
    return None


def calculate_time_difference(start_str, end_str):
    """Calculate time difference between start and end times."""
    start_dt = parse_datetime_string(start_str)
    end_dt = parse_datetime_string(end_str)
    
    if start_dt and end_dt:
        diff = end_dt - start_dt
        total_seconds = diff.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return "N/A"


def analyze_excel_file(file_path):
    """Analyze an Excel file and count pass/fail status from 2nd column (index 1)."""
    total_count = 0
    fail_count = 0
    pass_count = 0
    error_message = None
    execution_date = None
    execution_time = None
    
    # Try to read as .xls using xlrd first
    try:
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)  # Read first sheet
        
        # Check if sheet has at least 2 columns
        if sheet.ncols < 2:
            return {
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'total': 0,
                'fail': 0,
                'pass': 0,
                'execution_date': None,
                'execution_time': None,
                'error': 'File has less than 2 columns'
            }
        
        # Extract execution date and time from row 0 (0th row)
        try:
            row_0_values = sheet.row_values(0)
            # Execution date from column 2 (index 2)
            if len(row_0_values) > 2:
                date_val = row_0_values[2]
                if date_val:
                    # Check if it's an Excel date serial number
                    if isinstance(date_val, (int, float)):
                        try:
                            date_tuple = xlrd.xldate_as_tuple(date_val, workbook.datemode)
                            execution_date = datetime(*date_tuple).strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            execution_date = str(date_val).strip()
                    else:
                        execution_date = str(date_val).strip()
            # Start time from column 3 (index 3) and end time from column 4 (index 4)
            if len(row_0_values) > 4:
                start_val = row_0_values[3]
                end_val = row_0_values[4]
                if start_val and end_val:
                    # Convert Excel serial numbers to datetime strings if needed
                    start_time = None
                    end_time = None
                    if isinstance(start_val, (int, float)):
                        try:
                            start_tuple = xlrd.xldate_as_tuple(start_val, workbook.datemode)
                            start_time = datetime(*start_tuple).strftime("%Y-%m-%d-%H-%M-%S")
                        except:
                            start_time = str(start_val).strip()
                    else:
                        start_time = str(start_val).strip()
                    
                    if isinstance(end_val, (int, float)):
                        try:
                            end_tuple = xlrd.xldate_as_tuple(end_val, workbook.datemode)
                            end_time = datetime(*end_tuple).strftime("%Y-%m-%d-%H-%M-%S")
                        except:
                            end_time = str(end_val).strip()
                    else:
                        end_time = str(end_val).strip()
                    
                    if start_time and end_time:
                        execution_time = calculate_time_difference(start_time, end_time)
        except Exception as e:
            # If we can't parse the date/time, continue
            pass
        
        # Read from row 1 (skip header row 0) to end
        for row_idx in range(2, sheet.nrows):
            try:
                row_values = sheet.row_values(row_idx)
                if len(row_values) > 1:
                    status = str(row_values[1]).strip().lower()  # 2nd column (index 1)
                    
                    # Count only if status is not empty
                    if status:
                        total_count += 1
                        if status in ['fail', 'failure', 'failed', 'f']:
                            fail_count += 1
                        elif status in ['pass', 'passed', 'p']:
                            pass_count += 1
            except Exception as e:
                # Skip problematic rows
                continue
        
        return {
            'file_name': os.path.basename(file_path),
            'file_path': file_path,
            'total': total_count,
            'fail': fail_count,
            'pass': pass_count,
            'execution_date': execution_date,
            'execution_time': execution_time,
            'error': error_message
        }
    except xlrd.biffh.XLRDError as e:
        error_str = str(e)
        # If xlrd says it's an xlsx file, try openpyxl
        # Files with .xls extension might actually be .xlsx format
        if 'xlsx' in error_str.lower() and OPENPYXL_AVAILABLE:
            error_message = None
            try:
                # openpyxl checks file extension, so create temp copy with .xlsx extension
                temp_file = None
                try:
                    if not file_path.lower().endswith('.xlsx'):
                        # Create temporary file with .xlsx extension
                        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
                        os.close(temp_fd)
                        shutil.copy2(file_path, temp_file)
                        file_to_read = temp_file
                    else:
                        file_to_read = file_path
                    
                    # Try to read as xlsx
                    workbook = openpyxl.load_workbook(file_to_read, data_only=True, read_only=False)
                    sheet = workbook.active
                finally:
                    # Clean up temp file if created
                    if temp_file and os.path.exists(temp_file):
                        os.remove(temp_file)
                
                # Check if sheet has at least 2 columns
                if sheet.max_column < 2:
                    return {
                        'file_name': os.path.basename(file_path),
                        'file_path': file_path,
                        'total': 0,
                        'fail': 0,
                        'pass': 0,
                        'execution_date': None,
                        'execution_time': None,
                        'error': 'File has less than 2 columns'
                    }
                
                # Extract execution date and time from row 1 (0th row in openpyxl is row 1)
                try:
                    # Execution date from column 3 (index 2 -> column 3 in openpyxl)
                    if sheet.max_column >= 3:
                        execution_date_cell = sheet.cell(row=1, column=3).value
                        if execution_date_cell:
                            # openpyxl returns datetime objects when data_only=True
                            if isinstance(execution_date_cell, datetime):
                                execution_date = execution_date_cell.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                execution_date = str(execution_date_cell).strip()
                    # Start time from column 4 (index 3 -> column 4) and end time from column 5 (index 4 -> column 5)
                    if sheet.max_column >= 5:
                        start_time_cell = sheet.cell(row=1, column=4).value
                        end_time_cell = sheet.cell(row=1, column=5).value
                        start_time = None
                        end_time = None
                        if start_time_cell:
                            if isinstance(start_time_cell, datetime):
                                start_time = start_time_cell.strftime("%Y-%m-%d-%H-%M-%S")
                            else:
                                start_time = str(start_time_cell).strip()
                        if end_time_cell:
                            if isinstance(end_time_cell, datetime):
                                end_time = end_time_cell.strftime("%Y-%m-%d-%H-%M-%S")
                            else:
                                end_time = str(end_time_cell).strip()
                        if start_time and end_time:
                            execution_time = calculate_time_difference(start_time, end_time)
                except Exception as e:
                    # If we can't parse the date/time, continue
                    pass
                
                # Read from row 2 (skip header row 1) to end
                # openpyxl uses 1-based indexing
                for row_idx in range(2, sheet.max_row + 1):
                    try:
                        cell_value = sheet.cell(row=row_idx, column=2).value  # 2nd column
                        if cell_value is not None:
                            status = str(cell_value).strip().lower()
                            
                            # Count only if status is not empty
                            if status:
                                total_count += 1
                                if status in ['fail', 'failure', 'failed', 'f']:
                                    fail_count += 1
                                elif status in ['pass', 'passed', 'p']:
                                    pass_count += 1
                    except Exception as e:
                        # Skip problematic rows
                        continue
                
                workbook.close()
                return {
                    'file_name': os.path.basename(file_path),
                    'file_path': file_path,
                    'total': total_count,
                    'fail': fail_count,
                    'pass': pass_count,
                    'execution_date': execution_date,
                    'execution_time': execution_time,
                    'error': error_message
                }
            except Exception as e2:
                return {
                    'file_name': os.path.basename(file_path),
                    'file_path': file_path,
                    'total': 0,
                    'fail': 0,
                    'pass': 0,
                    'execution_date': None,
                    'execution_time': None,
                    'error': f"Error reading with openpyxl: {str(e2)}"
                }
        else:
            # xlrd failed for other reasons
            return {
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'total': 0,
                'fail': 0,
                'pass': 0,
                'execution_date': None,
                'execution_time': None,
                'error': f"Error reading with xlrd: {error_str}"
            }
    except Exception as e:
        # Catch any other exceptions
        return {
            'file_name': os.path.basename(file_path),
            'file_path': file_path,
            'total': 0,
            'fail': 0,
            'pass': 0,
            'execution_date': None,
            'execution_time': None,
            'error': f"Unexpected error: {str(e)}"
        }


def create_summary_excel(results, output_path):
    """Create a summary Excel file with the results."""
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet('Summary')
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#366092',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })
    
    number_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    fail_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#FFC7CE',
        'font_color': '#9C0006'
    })
    
    pass_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#C6EFCE',
        'font_color': '#006100'
    })
    
    error_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter',
        'bg_color': '#FFEB9C',
        'font_color': '#9C6500'
    })
    
    # Write headers
    headers = ['Excel File Name', 'File Path', 'Execution Date', 'Execution Time', 'Total Tests', 'Pass', 'Fail', 'Failure Rate (%)', 'Status']
    for col_idx, header in enumerate(headers):
        worksheet.write(0, col_idx, header, header_format)
    
    # Write data
    row_idx = 1
    for result in results:
        worksheet.write(row_idx, 0, result['file_name'], cell_format)
        worksheet.write(row_idx, 1, result['file_path'], cell_format)
        worksheet.write(row_idx, 2, result.get('execution_date', 'N/A'), cell_format)
        worksheet.write(row_idx, 3, result.get('execution_time', 'N/A'), cell_format)
        worksheet.write(row_idx, 4, result['total'], number_format)
        worksheet.write(row_idx, 5, result['pass'], pass_format if result['pass'] > 0 else number_format)
        worksheet.write(row_idx, 6, result['fail'], fail_format if result['fail'] > 0 else number_format)
        
        # Calculate failure rate
        if result['total'] > 0:
            failure_rate = (result['fail'] / result['total']) * 100
            worksheet.write(row_idx, 7, f"{failure_rate:.2f}%", number_format)
        else:
            worksheet.write(row_idx, 7, "N/A", number_format)
        
        # Overall status
        if result['error']:
            worksheet.write(row_idx, 8, "Error", error_format)
        elif result['total'] == 0:
            worksheet.write(row_idx, 8, "No Data", error_format)
        elif result['fail'] == 0:
            worksheet.write(row_idx, 8, "All Pass", pass_format)
        else:
            worksheet.write(row_idx, 8, "Has Failures", fail_format)
        
        # # Error message if any
        # if result['error']:
        #     worksheet.write(row_idx, 7, result['error'], error_format)
        # else:
        #     worksheet.write(row_idx, 7, "", cell_format)
        #
        row_idx += 1
    
    # Add summary row at the end
    total_files = len(results)
    total_tests = sum(r['total'] for r in results)
    total_passes = sum(r['pass'] for r in results)
    total_fails = sum(r['fail'] for r in results)
    
    summary_row = row_idx + 1
    worksheet.write(summary_row, 0, "TOTAL SUMMARY", header_format)
    worksheet.write(summary_row, 1, "", header_format)
    worksheet.write(summary_row, 2, "", header_format)  # Execution Date
    worksheet.write(summary_row, 3, "", header_format)  # Execution Time
    worksheet.write(summary_row, 4, total_tests, header_format)
    worksheet.write(summary_row, 5, total_passes, pass_format if total_passes > 0 else header_format)
    worksheet.write(summary_row, 6, total_fails, fail_format if total_fails > 0 else header_format)
    
    if total_tests > 0:
        overall_failure_rate = (total_fails / total_tests) * 100
        worksheet.write(summary_row, 7, f"{overall_failure_rate:.2f}%", header_format)
    else:
        worksheet.write(summary_row, 7, "N/A", header_format)
    
    worksheet.write(summary_row, 8, f"Files: {total_files}", header_format)
    
    # Set column widths
    worksheet.set_column(0, 0, 40)  # File Name
    worksheet.set_column(1, 1, 60)  # File Path
    worksheet.set_column(2, 2, 20)  # Execution Date
    worksheet.set_column(3, 3, 15)  # Execution Time
    worksheet.set_column(4, 4, 12)  # Total Tests
    worksheet.set_column(5, 5, 10)  # Pass
    worksheet.set_column(6, 6, 10)  # Fail
    worksheet.set_column(7, 7, 15)  # Failure Rate
    worksheet.set_column(8, 8, 15)  # Status
    
    workbook.close()
    print(f"Summary Excel file created: {output_path}")


def main():
    """Main function to process all Excel files and create summary."""
    # Define paths
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # input_directory = os.path.join(base_dir, 'PythonWorkingScripts_Output')
    input_directory = '/Users/senthil/Desktop/sprint wise /213'
    output_directory = os.path.join(base_dir, 'PythonWorkingScripts_Output')
    
    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    output_filename = f"Excel_Reports_Summary_{timestamp}.xls"
    output_path = os.path.join(output_directory, output_filename)
    
    print(f"Searching for Excel files in: {input_directory}")
    
    # Find all Excel files
    excel_files = find_all_excel_files(input_directory)
    print(f"Found {len(excel_files)} Excel file(s)")
    
    if not excel_files:
        print("No Excel files found!")
        return
    
    # Analyze each Excel file
    results = []
    for excel_file in excel_files:
        print(f"Processing: {os.path.basename(excel_file)}")
        result = analyze_excel_file(excel_file)
        results.append(result)
        print(f"  - Total: {result['total']}, Pass: {result['pass']}, Fail: {result['fail']}")
        if result['error']:
            print(f"  - Error: {result['error']}")
    
    # Create summary Excel file
    print(f"\nCreating summary Excel file...")
    create_summary_excel(results, output_path)
    
    # Print summary statistics
    total_tests = sum(r['total'] for r in results)
    total_passes = sum(r['pass'] for r in results)
    total_fails = sum(r['fail'] for r in results)
    
    print(f"\n=== SUMMARY ===")
    print(f"Total Excel files processed: {len(excel_files)}")
    print(f"Total tests: {total_tests}")
    print(f"Total passes: {total_passes}")
    print(f"Total failures: {total_fails}")
    if total_tests > 0:
        print(f"Overall failure rate: {(total_fails/total_tests)*100:.2f}%")
    print(f"\nSummary file saved to: {output_path}")


if __name__ == "__main__":
    main()

